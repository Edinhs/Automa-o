import csv
import io
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, time, timedelta, timezone
from html import escape
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import report_delivery_dir, runtime_path, settings
from app.core.report_i18n import (
    DEFAULT_LANGUAGE,
    card_strings,
    misc_string,
    normalize_language,
    poster_labels,
    poster_strings,
    report_block_title,
    report_headers,
    report_type_label,
    simplificado_observation,
    simplificado_status,
    summary_labels,
)
from app.core.report_i18n import status_label as i18n_status_label
from app.core.timezone import app_timezone, now_sao_paulo_naive, sao_paulo_utc_iso, to_sao_paulo_naive
from app.db.session import get_db
from app.models.automation import Automation
from app.models.execution import ExecutionLog, ExecutionReport
from app.models.file import WorkspaceFile
from app.models.workspace import Workspace
from app.models.agent import AgentTask
from app.models.schedule import Schedule
from app.routers.executions import (
    STATUS_FILTERS,
    files_for_task,
    file_status_counts,
    finished_at,
    group_files,
    group_status,
    group_tasks_by_origin,
    normalize_text,
    task_automation_id,
    task_workspace_id,
)
from app.services.audit import create_log
from app.services.access_requests import get_engineers_count

router = APIRouter()

REPORT_SOURCE_SCOPE = "folder_monitoring_detection"
DETECTION_SOURCE = "folder_monitoring"
LOCAL_REPORT_EVENTS = {
    "folder_not_found",
    "folder_inaccessible",
    "folder_scan_failed",
    "file_signature_failed",
    "subfolder_inaccessible",
    "item_inaccessible",
    "copy_failed",
    "no_files_copied",
}
REPORT_BLOCKS = {
    "files": "Arquivos Detectados",
    "local_errors": "Erros Locais",
    "automations": "Automações",
    "updated_files": "Arquivos Atualizados",
    "workspaces": "Workspaces",
    "schedules": "Agendamentos",
    "executions": "Execuções",
    "simplificado": "Relatório Simplificado",
}
REPORT_TYPES = {
    "relatorio geral": ("Relatório Geral", ["files", "local_errors", "automations", "updated_files", "workspaces", "schedules", "executions", "simplificado"]),
    "relatorio arquivos": ("Relatório Arquivos", ["files"]),
    "relatorio erros locais": ("Relatório Erros Locais", ["local_errors"]),
    "relatorio automacao": ("Relatório Automação", ["automations"]),
    "relatorio atualizados": ("Relatório Atualizados", ["updated_files"]),
    "relatorio workspace": ("Relatório Workspace", ["workspaces"]),
    "relatorio agendamento": ("Relatório Agendamento", ["schedules"]),
    "relatorio execucoes": ("Relatório Execuções", ["executions"]),
    "relatorio simplificado": ("Relatório Simplificado", ["simplificado"]),
}
MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
    "csv": "text/csv; charset=utf-8",
    "json": "application/json; charset=utf-8",
}


@dataclass
class ReportSection:
    key: str
    title: str
    headers: list[str]
    rows: list[list[Any]]


def normalize_key(value: str | None) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().strip().split())


def parse_report_type(value: str | None) -> str:
    raw = value or "Relatório Geral"
    requested = raw.split("|", 1)[0].strip() or "Relatório Geral"
    report_definition = REPORT_TYPES.get(normalize_key(requested))
    if not report_definition:
        raise HTTPException(422, detail="Tipo de relatorio invalido. Use apenas Geral, Arquivos, Erros Locais, Automação, Atualizados, Workspace, Agendamento, Execuções ou Simplificado.")
    return report_definition[0]


def parse_file_format(value: str | None, fallback_type: str | None = None) -> str:
    raw = (value or "").strip().lower().lstrip(".")
    if raw:
        return "xlsx" if raw in {"excel", "xls"} else raw
    parts = (fallback_type or "").split("|")
    if len(parts) > 1:
        return parse_file_format(parts[-1])
    return "csv"


def parse_environment_mode(value: Any) -> str:
    normalized = normalize_key(str(value or ""))
    if normalized in {"developer", "desenvolvedor", "dev", "test", "teste"}:
        return "developer"
    return "operational"


def parse_dt(value: Any, end_of_day: bool = False) -> datetime | None:
    """Interpreta a data informada (date picker = horario de Sao Paulo) e devolve o limite em
    UTC NAIVE, para comparar corretamente com as colunas gravadas em UTC (created_at/detected_at/
    started_at = datetime.utcnow). Antes, um filtro de dia LOCAL era comparado contra timestamps
    UTC, gerando ate 3h de erro nas bordas do dia (inclusao/exclusao indevida de arquivos)."""
    if not value:
        return None
    tz = app_timezone()
    if isinstance(value, datetime):
        local = value.astimezone(tz).replace(tzinfo=None) if value.tzinfo else value
    else:
        raw = str(value).strip().replace("Z", "+00:00")
        try:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
                local = datetime.combine(datetime.fromisoformat(raw).date(), time.max if end_of_day else time.min)
            else:
                parsed = datetime.fromisoformat(raw)
                local = parsed.astimezone(tz).replace(tzinfo=None) if parsed.tzinfo else parsed
        except ValueError:
            return None
    # Sao Paulo local -> UTC naive (mesma convencao das colunas comparadas em within_period).
    return local.replace(tzinfo=tz).astimezone(timezone.utc).replace(tzinfo=None)


def safe_int(value: Any) -> int | None:
    if value in [None, "", "Todos"]:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def format_dt(value: datetime | None) -> str:
    return value.strftime("%d/%m/%Y %H:%M:%S") if value else ""


def fmt_utc(value: datetime | None) -> str:
    """Formata uma data gravada em UTC convertendo para o horario de Sao Paulo (exibicao fiel
    ao relogio local do usuario, consistente com o resto do app). Use em colunas UTC
    (created_at/detected_at/started_at...). Colunas ja locais (next_run_at/last_run_at) usam
    format_dt direto."""
    return format_dt(to_sao_paulo_naive(value, assume_utc=True))


def parse_json(raw: str | None) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def within_period(value: datetime | None, start: datetime | None, end: datetime | None) -> bool:
    if not value:
        return True
    if start and value < start:
        return False
    if end and value > end:
        return False
    return True


def filters_from_payload(data: dict[str, Any]) -> dict[str, Any]:
    return {
        "start": parse_dt(data.get("date_start"), end_of_day=False),
        "end": parse_dt(data.get("date_end"), end_of_day=True),
        "automation_id": safe_int(data.get("automation_id")),
        "workspace_id": safe_int(data.get("workspace_id")),
        "status": None if data.get("status") in [None, "", "Todos"] else str(data.get("status")),
        "source_task_id": safe_int(data.get("source_task_id")),
    }


def sections_for_type(report_type: str) -> list[str]:
    return REPORT_TYPES[normalize_key(parse_report_type(report_type))][1]


def clean_filename(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "relatorio")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("_")
    return text.lower() or "relatorio"


def report_filename(report_type: str, file_format: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S_%f")
    return f"{clean_filename(report_type)}_{timestamp}.{file_format}"


def lookup_maps(db: Session) -> tuple[dict[int, str], dict[int, str]]:
    automations = {
        item.id: item.name
        for item in db.query(Automation).filter(Automation.is_deleted == False, Automation.type == DETECTION_SOURCE).all()
    }
    workspaces = {
        item.id: item.name
        for item in db.query(Workspace).filter(Workspace.is_deleted == False).all()
    }
    return automations, workspaces


def reportable_automation_ids(db: Session) -> set[int]:
    return {
        item.id
        for item in db.query(Automation).filter(Automation.is_deleted == False, Automation.type == DETECTION_SOURCE).all()
    }


def block_files(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    automation_names, workspace_names = names
    permitted_automations = reportable_automation_ids(db)
    query = db.query(WorkspaceFile).filter(
        WorkspaceFile.is_deleted == False,
        WorkspaceFile.detection_source == DETECTION_SOURCE,
        WorkspaceFile.detection_task_id.isnot(None),
    )
    if filters["source_task_id"]:
        query = query.filter(WorkspaceFile.detection_task_id == filters["source_task_id"])
    rows = []
    for item in query.order_by(WorkspaceFile.detected_at.desc(), WorkspaceFile.id.desc()).all():
        if item.automation_id not in permitted_automations:
            continue
        if item.detection_classification not in {"new", "updated", "audit_duplicate"}:
            continue
        if filters["automation_id"] and item.automation_id != filters["automation_id"]:
            continue
        if filters["workspace_id"] and item.workspace_id != filters["workspace_id"]:
            continue
        if filters["status"] and item.detection_classification.lower() != filters["status"].lower():
            continue
        event_date = item.detected_at or item.created_at
        if not within_period(event_date, filters["start"], filters["end"]):
            continue
        rows.append([
            item.id,
            item.file_name or "",
            automation_names.get(item.automation_id, item.automation_id or ""),
            workspace_names.get(item.workspace_id, item.workspace_id or ""),
            item.detection_classification,
            item.extension or "",
            item.size_bytes or "",
            item.original_path or "",
            fmt_utc(event_date),
            item.detection_task_id,
        ])
    return ReportSection(
        "files",
        report_block_title("files", language),
        report_headers("files", language),
        rows,
    )


def block_local_errors(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    automation_names, _ = names
    permitted_automations = reportable_automation_ids(db)
    query = db.query(ExecutionLog).filter(ExecutionLog.level == "error")
    if filters["source_task_id"]:
        query = query.filter(ExecutionLog.task_id == filters["source_task_id"])
    rows = []
    for log in query.order_by(ExecutionLog.created_at.desc(), ExecutionLog.id.desc()).all():
        metadata = parse_json(log.metadata_json)
        if metadata.get("report_source") != REPORT_SOURCE_SCOPE:
            continue
        event = metadata.get("report_event")
        if event not in LOCAL_REPORT_EVENTS:
            continue
        if log.automation_id not in permitted_automations:
            continue
        if filters["automation_id"] and log.automation_id != filters["automation_id"]:
            continue
        if not within_period(log.created_at, filters["start"], filters["end"]):
            continue
        rows.append([
            log.id,
            event,
            automation_names.get(log.automation_id, log.automation_id or ""),
            log.task_id or "",
            log.message or "",
            fmt_utc(log.created_at),
            json.dumps({key: value for key, value in metadata.items() if key not in {"report_source", "report_event"}}, ensure_ascii=False),
        ])
    return ReportSection(
        "local_errors",
        report_block_title("local_errors", language),
        report_headers("local_errors", language),
        rows,
    )


def block_automations(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    _, workspace_names = names
    query = db.query(Automation).filter(Automation.is_deleted == False)
    if filters["automation_id"]:
        query = query.filter(Automation.id == filters["automation_id"])
    rows = []
    for item in query.order_by(Automation.id.desc()).all():
        if not within_period(item.created_at, filters["start"], filters["end"]):
            continue
        rows.append([
            item.id,
            item.name or "",
            item.description or "",
            item.type or "",
            item.status or "",
            item.folder_path or "",
            item.temp_folder_path or "",
            fmt_utc(item.created_at),
        ])
    return ReportSection("automations", report_block_title("automations", language), report_headers("automations", language), rows)


def block_updated_files(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    automation_names, workspace_names = names
    permitted_automations = reportable_automation_ids(db)
    query = db.query(WorkspaceFile).filter(
        WorkspaceFile.is_deleted == False,
        WorkspaceFile.detection_source == DETECTION_SOURCE,
        WorkspaceFile.detection_task_id.isnot(None),
        WorkspaceFile.detection_classification == "updated",
    )
    if filters["source_task_id"]:
        query = query.filter(WorkspaceFile.detection_task_id == filters["source_task_id"])
    rows = []
    for item in query.order_by(WorkspaceFile.detected_at.desc(), WorkspaceFile.id.desc()).all():
        if item.automation_id not in permitted_automations:
            continue
        if filters["automation_id"] and item.automation_id != filters["automation_id"]:
            continue
        if filters["workspace_id"] and item.workspace_id != filters["workspace_id"]:
            continue
        event_date = item.detected_at or item.created_at
        if not within_period(event_date, filters["start"], filters["end"]):
            continue
        rows.append([
            item.id,
            item.file_name or "",
            automation_names.get(item.automation_id, item.automation_id or ""),
            workspace_names.get(item.workspace_id, item.workspace_id or ""),
            item.detection_classification,
            item.extension or "",
            item.size_bytes or "",
            item.original_path or "",
            fmt_utc(event_date),
            item.detection_task_id,
        ])
    return ReportSection("updated_files", report_block_title("updated_files", language), report_headers("updated_files", language), rows)


def block_workspaces(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    query = db.query(Workspace).filter(Workspace.is_deleted == False)
    if filters["workspace_id"]:
        query = query.filter(Workspace.id == filters["workspace_id"])
    rows = []
    for item in query.order_by(Workspace.id.desc()).all():
        if not within_period(item.created_at, filters["start"], filters["end"]):
            continue
        rows.append([
            item.id,
            item.name or "",
            item.description or "",
            item.playground_workspace_id or "",
            item.playground_url or "",
            item.embedding_model or "",
            item.data_languages or "",
            item.status or "",
            item.created_via or "",
            fmt_utc(item.created_at),
        ])
    return ReportSection("workspaces", report_block_title("workspaces", language), report_headers("workspaces", language), rows)


def block_schedules(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    automation_names, _ = names
    query = db.query(Schedule).filter(Schedule.is_deleted == False)
    if filters["automation_id"]:
        query = query.filter(Schedule.automation_id == filters["automation_id"])
    rows = []
    for item in query.order_by(Schedule.id.desc()).all():
        if not within_period(item.created_at, filters["start"], filters["end"]):
            continue
        rows.append([
            item.id,
            item.name or "",
            automation_names.get(item.automation_id, item.automation_id or ""),
            item.frequency_type or "",
            item.time_of_day or "",
            item.days_of_week or "",
            item.day_of_month or "",
            format_dt(item.next_run_at),
            format_dt(item.last_run_at),
            item.status or "",
            fmt_utc(item.created_at),
        ])
    return ReportSection("schedules", report_block_title("schedules", language), report_headers("schedules", language), rows)


def block_executions(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    automation_names, workspace_names = names
    # Uma "execucao" = 1 INICIALIZACAO da automacao (nao 1 task). As tasks satelites do mesmo run
    # (o reenvio de PDF pos-monitoramento, que tambem e upload_files_to_workspace e carrega
    # origin_task_id) sao AGRUPADAS na task raiz -- exatamente como list_executions em
    # routers/executions.py, para o relatorio bater com o Historico ao vivo (1 run = 1 linha
    # agregada, sem inflar). Buscamos todas as upload_files_to_workspace, agrupamos por origem e
    # so entao filtramos/agregamos no nivel do GRUPO.
    all_tasks = (
        db.query(AgentTask)
        .filter(
            AgentTask.is_deleted == False,
            AgentTask.started_at.isnot(None),
            AgentTask.task_type == "upload_files_to_workspace",
        )
        .order_by(AgentTask.started_at.desc(), AgentTask.id.desc())
        .all()
    )
    groups = group_tasks_by_origin(all_tasks)
    status_target = (
        STATUS_FILTERS.get(normalize_text(filters["status"]), filters["status"])
        if filters["status"]
        else None
    )
    ordered_roots = sorted(
        groups.keys(),
        key=lambda rid: max((t.started_at for t in groups[rid] if t.started_at), default=datetime.min),
        reverse=True,
    )
    rows = []
    for root_id in ordered_roots:
        group = groups[root_id]
        root = min(group, key=lambda t: t.id)
        payload = parse_json(root.payload_json)
        automation_id = task_automation_id(root, payload)
        workspace_id = task_workspace_id(root, payload)

        # source_task_id pode ser a raiz OU uma task satelite do run -> casa com o grupo inteiro.
        if filters["source_task_id"] and filters["source_task_id"] not in {t.id for t in group}:
            continue
        if filters["automation_id"] and automation_id != filters["automation_id"]:
            continue
        if filters["workspace_id"] and workspace_id != filters["workspace_id"]:
            continue
        group_started = min((t.started_at for t in group if t.started_at), default=root.started_at)
        if not within_period(group_started, filters["start"], filters["end"]):
            continue
        status_key = group_status(group) if len(group) > 1 else (root.status or "pending")
        if status_target and status_key != status_target:
            continue

        files = group_files(db, group) if len(group) > 1 else files_for_task(db, root, payload)
        counts = file_status_counts(files)
        finished_candidates = [f for f in (finished_at(t) for t in group) if f]
        group_finished = (
            max(finished_candidates)
            if finished_candidates and len(finished_candidates) == len(group)
            else None
        )
        end = group_finished or datetime.utcnow()
        duration = max(int((end - group_started).total_seconds()), 0) if group_started else 0
        status_label = i18n_status_label(status_key, language)

        rows.append([
            root.id,
            root.task_type or "",
            automation_names.get(automation_id, automation_id or ""),
            workspace_names.get(workspace_id, workspace_id or payload.get("workspace_name") or ""),
            fmt_utc(group_started),
            fmt_utc(group_finished),
            duration,
            counts["total"],
            counts["success"],
            counts["errors"],
            status_label,
        ])
    return ReportSection(
        "executions",
        report_block_title("executions", language),
        report_headers("executions", language),
        rows,
    )


def block_simplificado(db: Session, filters: dict[str, Any], names: tuple[dict[int, str], dict[int, str]], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    query = db.query(Workspace).filter(Workspace.is_deleted == False)
    if filters["workspace_id"]:
        query = query.filter(Workspace.id == filters["workspace_id"])
    workspaces = [
        ws for ws in query.order_by(Workspace.id.asc()).all()
        if within_period(ws.created_at, filters["start"], filters["end"])
    ]
    workspace_ids = [ws.id for ws in workspaces]

    # Carrega todos os arquivos ativos dos workspaces em uma unica query (evita N+1).
    files_by_ws: dict[int, list[WorkspaceFile]] = {ws_id: [] for ws_id in workspace_ids}
    if workspace_ids:
        for f in db.query(WorkspaceFile).filter(
            WorkspaceFile.is_deleted == False,
            WorkspaceFile.workspace_id.in_(workspace_ids),
        ).all():
            files_by_ws.setdefault(f.workspace_id, []).append(f)

    rows = []
    for ws in workspaces:
        files = files_by_ws.get(ws.id, [])
        total = len(files)

        def is_ready(f: WorkspaceFile) -> bool:
            return (f.status or "").lower() == "ready" or (f.playground_status or "").lower() == "ready"

        def is_error(f: WorkspaceFile) -> bool:
            return (f.status or "").lower() in {"failed", "manual_review", "pending_retry"}

        def is_sent(f: WorkspaceFile) -> bool:
            return (f.status or "").lower() in {"uploaded", "ready"}

        # status_key e um TOKEN canonico (COMPLETO/PROGRESSO/ERRO); a regra de negocio do card
        # (compute_card_business / build_card_summary) le esse token em pt. O valor exibido na linha
        # e localizado (idem OBSERVAÇÃO) sem alterar o token canonico.
        if total == 0:
            percentage, status_key, obs_key = "10%", "PROGRESSO", "created"
        else:
            all_ready = all(is_ready(f) for f in files)
            any_error = any(is_error(f) for f in files)
            if any_error and not all_ready:
                percentage, status_key, obs_key = "90%", "ERRO", "error_handling"
            elif all_ready:
                percentage, status_key, obs_key = "100%", "COMPLETO", "available"
            elif all(is_sent(f) for f in files):
                percentage, status_key, obs_key = "70%", "PROGRESSO", "files_sent"
            elif any(is_sent(f) for f in files):
                percentage, status_key, obs_key = "40%", "PROGRESSO", "sending"
            else:
                percentage, status_key, obs_key = "40%", "PROGRESSO", "sending"

        rows.append([
            ws.name or "",
            percentage,
            simplificado_status(status_key, language),
            simplificado_observation(obs_key, language),
            fmt_utc(ws.updated_at),
            total,
        ])

    return ReportSection(
        "simplificado",
        report_block_title("simplificado", language),
        report_headers("simplificado", language),
        rows,
    )


BLOCK_BUILDERS = {
    "files": block_files,
    "local_errors": block_local_errors,
    "automations": block_automations,
    "updated_files": block_updated_files,
    "workspaces": block_workspaces,
    "schedules": block_schedules,
    "executions": block_executions,
    "simplificado": block_simplificado,
}


def build_sections(db: Session, report_type: str, filters: dict[str, Any], language: str = DEFAULT_LANGUAGE) -> list[ReportSection]:
    names = lookup_maps(db)
    return [BLOCK_BUILDERS[key](db, filters, names, language) for key in sections_for_type(report_type)]


def summary_section(report_type: str, file_format: str, filters: dict[str, Any], sections: list[ReportSection], language: str = DEFAULT_LANGUAGE) -> ReportSection:
    lb = summary_labels(language)
    all_label = lb["all"]
    rows = [
        [lb["type"], report_type_label(report_type, language)],
        [lb["format"], file_format.upper()],
        [lb["source"], lb["source_value"]],
        [lb["date_start"], fmt_utc(filters["start"])],
        [lb["date_end"], fmt_utc(filters["end"])],
        [lb["automation"], filters["automation_id"] or all_label],
        [lb["workspace"], filters["workspace_id"] or all_label],
        [lb["classification"], filters["status"] or all_label],
        [lb["cycle"], filters["source_task_id"] or all_label],
        [lb["generated_at"], format_dt(now_sao_paulo_naive())],
    ]
    rows.extend([[section.title, len(section.rows)] for section in sections])
    return ReportSection("summary", lb["title"], list(lb["headers"]), rows)


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return format_dt(value)
    return str(value)


def safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]+", " ", value).strip()[:31] or "Aba"
    name = base
    index = 2
    while name in used:
        suffix = f" {index}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(name)
    return name


def build_excel(report_type: str, file_format: str, filters: dict[str, Any], sections: list[ReportSection], language: str = DEFAULT_LANGUAGE) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    used_names: set[str] = set()
    all_sections = [summary_section(report_type, file_format, filters, sections, language), *sections]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for index, section in enumerate(all_sections):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = safe_sheet_name(section.title, used_names)
        for col_idx, header in enumerate(section.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(section.rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=normalize_cell(value))
        if not section.rows:
            ws.cell(row=2, column=1, value=misc_string("no_records", language))
        for column in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in column), default=10)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 4, 12), 70)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def pdf_table(section: ReportSection, styles, language: str = DEFAULT_LANGUAGE):
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    data_rows = section.rows or [[misc_string("no_records", language)] + [""] * (len(section.headers) - 1)]
    table_data = [
        [Paragraph(f"<b>{escape(str(header))}</b>", styles["BodyText"]) for header in section.headers],
        *[[Paragraph(escape(normalize_cell(value)), styles["BodyText"]) for value in row] for row in data_rows[:80]],
    ]
    col_width = 760 / max(len(section.headers), 1)
    table = Table(table_data, colWidths=[col_width] * len(section.headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def build_pdf(report_type: str, file_format: str, filters: dict[str, Any], sections: list[ReportSection], language: str = DEFAULT_LANGUAGE) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=cm, rightMargin=cm, topMargin=cm, bottomMargin=cm)
    styles = getSampleStyleSheet()
    styles["BodyText"].fontSize = 6
    styles["BodyText"].leading = 7
    elements = [
        Paragraph("<b>Stellantis Automation HUB</b>", styles["Title"]),
        Paragraph(escape(misc_string("pdf_report_heading", language).format(type=report_type_label(report_type, language))), styles["Heading2"]),
        Spacer(1, 0.3 * cm),
        pdf_table(summary_section(report_type, file_format, filters, sections, language), styles, language),
        Spacer(1, 0.5 * cm),
    ]
    for index, section in enumerate(sections):
        if index:
            elements.append(PageBreak())
        elements.append(Paragraph(escape(section.title), styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * cm))
        elements.append(pdf_table(section, styles, language))
        if len(section.rows) > 80:
            elements.append(Spacer(1, 0.2 * cm))
            elements.append(Paragraph(escape(misc_string("pdf_showing", language).format(n=len(section.rows))), styles["Normal"]))
    doc.build(elements)
    return buf.getvalue()


def build_csv(report_type: str, file_format: str, filters: dict[str, Any], sections: list[ReportSection], language: str = DEFAULT_LANGUAGE) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    if normalize_key(report_type) == "relatorio geral":
        for section in [summary_section(report_type, file_format, filters, sections, language), *sections]:
            writer.writerow([section.title])
            writer.writerow(section.headers)
            writer.writerows(section.rows)
            writer.writerow([])
    else:
        section = sections[0]
        writer.writerow(section.headers)
        writer.writerows(section.rows)
    return buf.getvalue()


def build_json(report_type: str, file_format: str, filters: dict[str, Any], sections: list[ReportSection], language: str = DEFAULT_LANGUAGE) -> str:
    all_sections = [summary_section(report_type, file_format, filters, sections, language), *sections]
    payload = {
        "report_type": report_type,
        "language": language,
        "file_format": file_format,
        "generated_at": now_sao_paulo_naive().isoformat(timespec="seconds"),
        "period": {
            "start": fmt_utc(filters["start"]),
            "end": fmt_utc(filters["end"]),
        },
        "sections": [
            {
                "key": section.key,
                "title": section.title,
                "headers": list(section.headers),
                "rows": [
                    {header: normalize_cell(value) for header, value in zip(section.headers, row)}
                    for row in section.rows
                ],
            }
            for section in all_sections
        ],
    }
    return json.dumps(payload, ensure_ascii=False, default=str, indent=2)


def build_report_content(report_type: str, file_format: str, filters: dict[str, Any], db: Session, language: str = DEFAULT_LANGUAGE) -> bytes:
    language = normalize_language(language)
    sections = build_sections(db, report_type, filters, language)
    if file_format == "xlsx":
        return build_excel(report_type, file_format, filters, sections, language)
    if file_format == "pdf":
        return build_pdf(report_type, file_format, filters, sections, language)
    if file_format == "json":
        return build_json(report_type, file_format, filters, sections, language).encode("utf-8")
    return build_csv(report_type, file_format, filters, sections, language).encode("utf-8-sig")


def write_report_file(report_type: str, file_format: str, filters: dict[str, Any], db: Session, subfolder: str = None, language: str = DEFAULT_LANGUAGE) -> Path:
    reports_dir = runtime_path("REPORTS_PATH")
    if subfolder:
        reports_dir = reports_dir / subfolder
    reports_dir.mkdir(parents=True, exist_ok=True)
    path = reports_dir / report_filename(report_type, file_format)
    path.write_bytes(build_report_content(report_type, file_format, filters, db, language))
    return path


def report_type_and_format(rep: ExecutionReport) -> tuple[str, str]:
    report_type = parse_report_type(rep.type)
    return report_type, parse_file_format(None, rep.type)


def report_out(rep: ExecutionReport) -> dict[str, Any]:
    report_type, file_format = report_type_and_format(rep)
    return {
        "id": rep.id,
        "name": rep.name,
        "file_name": Path(rep.file_path).name if rep.file_path else rep.name,
        "report_type": report_type,
        "type": rep.type,
        "file_format": file_format,
        "language": normalize_language(getattr(rep, "language", None)),
        "file_path": rep.file_path,
        "status": rep.status,
        "source_scope": rep.source_scope,
        "generation_trigger": rep.generation_trigger,
        "source_task_id": rep.source_task_id,
        "period_start": rep.period_start,
        "period_end": rep.period_end,
        "generated_by_id": rep.generated_by_id,
        "generated_at": rep.created_at,
        "created_at": rep.created_at,
        "updated_at": rep.updated_at,
    }


def persist_report(
    db: Session,
    report_type: str,
    file_format: str,
    filters: dict[str, Any],
    generated_by_id: int | None,
    generation_trigger: str,
    source_task_id: int | None,
    deliver_to_folder: bool = False,
    language: str = DEFAULT_LANGUAGE,
) -> ExecutionReport:
    language = normalize_language(language)
    subfolder = "agendados" if generation_trigger == "automatic" else None
    path = write_report_file(report_type, file_format, filters, db, subfolder=subfolder, language=language)
    report = ExecutionReport(
        name=f"{report_type} ({file_format.upper()}) - {now_sao_paulo_naive():%d/%m/%Y %H:%M}",
        type=f"{report_type}|{file_format}",
        status="ready",
        file_path=str(path),
        source_scope=REPORT_SOURCE_SCOPE,
        generation_trigger=generation_trigger,
        source_task_id=source_task_id,
        period_start=filters["start"],
        period_end=filters["end"],
        generated_by_id=generated_by_id,
        language=language,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    create_log(
        db,
        "info",
        f"Report created: {report.name}",
        "report",
        report.id,
        metadata={"source_scope": REPORT_SOURCE_SCOPE, "generation_trigger": generation_trigger, "source_task_id": source_task_id},
    )
    # Copia para a pasta de entrega (REPORT_DELIVERY_PATH / Power Automate) SOMENTE para o
    # "Relatório Simplificado": geração manual entrega automaticamente; agendamento respeita o
    # toggle deliver_to_folder. Qualquer outro tipo segue o caminho natural (só REPORTS_PATH).
    should_deliver = report_type == "Relatório Simplificado" and (deliver_to_folder or generation_trigger == "manual")
    if should_deliver:
        try:
            bundle = report_delivery_bundle(db, report.id)
            delivery_path = write_report_to_delivery_folder(bundle)
            if delivery_path is not None:
                create_log(
                    db,
                    "info",
                    "Report copied to delivery folder",
                    "report",
                    report.id,
                    metadata={"delivery_path": str(delivery_path)},
                )
        except Exception as exc:
            create_log(
                db,
                "warning",
                f"Report delivery folder copy failed: {exc}",
                "report",
                report.id,
                metadata={"source_scope": REPORT_SOURCE_SCOPE},
            )
    return report


def filters_for_report(rep: ExecutionReport) -> dict[str, Any]:
    """Reconstroi o dict de filtros a partir de um ExecutionReport persistido."""
    return {
        "start": rep.period_start,
        "end": rep.period_end,
        "automation_id": None,
        "workspace_id": None,
        "status": None,
        "source_task_id": rep.source_task_id if rep.generation_trigger == "automatic" else None,
    }


def fallback_content(rep: ExecutionReport, file_format: str, db: Session) -> bytes:
    report_type, _ = report_type_and_format(rep)
    # Regenera no MESMO idioma persistido do relatorio (re-download/fallback consistente).
    return build_report_content(report_type, file_format, filters_for_report(rep), db, normalize_language(getattr(rep, "language", None)))


CARD_PREVIEW_ROW_LIMIT = 5
# Relatorio Simplificado: mostra TODOS os workspaces na previa (1 linha por workspace),
# com um teto de seguranca para nao estourar o limite de ~28 KB do Adaptive Card no Teams.
SIMPLIFICADO_PREVIEW_MAX = 100

# URL-placeholder do botao "Baixar PDF" no Adaptive Card. O fluxo do Power Automate
# substitui esta string EXATA pelo link de compartilhamento real do PDF (replace()).
DOWNLOAD_URL_PLACEHOLDER = "https://hub-report-download.invalid"

# URL-placeholder da IMAGEM (PNG) no card-imagem. O Power Automate substitui esta string EXATA
# pelo link direto (bytes) da imagem gerado pelo fluxo (ver GUIA_POWER_AUTOMATE.md, Parte I).
IMAGE_URL_PLACEHOLDER = "https://hub-report-image.invalid"


def _section_by_key(sections: list[ReportSection], key: str) -> ReportSection | None:
    return next((section for section in sections if section.key == key), None)


# Arquivos que ja foram "encontrados e enviados ao workspace" (o trabalho de setup que a HUB poupou).
SENT_FILE_STATUSES = ("uploaded", "ready", "resolved")

# Convite (manchete) do card semanal: o gancho e devolver tempo ao engenheiro, nao "ver o status".
CARD_HEADLINE = "🚀 Seu ambiente já está pronto — entre e crie seu agente"
CARD_INVITE_BODY = (
    "Esqueça baixar a SPEC, subir no workspace seguro e montar o ambiente: a automação já fez tudo "
    "isso. Entre no Playground e vá direto ao que importa — criar o agente no workspace do seu projeto."
)
# "Como pedir acesso" (parte do convite). O botao "Solicitar acesso" so aparece se REPORT_CARD_ACCESS_URL
# estiver setado; esta linha funciona como instrucao mesmo sem o botao.
CARD_ACCESS_LINE = '→ Não tem acesso ao workspace? Toque em "Solicitar acesso" abaixo e preencha o formulário.'
# Saude em 1 linha: previsao de correcao exibida quando ha itens em tratamento. Frase editavel
# (decisao de produto: nao ha campo de ETA por item — e uma promessa de SLA operacional do time).
CARD_HEALTH_ETA = "em até 1 dia útil"


def _format_hours(minutes: float) -> str:
    """Formata minutos economizados como horas legiveis (pt-BR: virgula decimal) para o card.

    < 1h -> minutos; 1-10h -> 1 casa com virgula (ex.: '1,5 h'); >= 10h -> sem decimal ('12 h').
    """
    minutes = max(0.0, float(minutes or 0))
    hours = minutes / 60.0
    if hours >= 10:
        return f"{hours:.0f} h"
    if hours >= 1:
        return f"{hours:.1f} h".replace(".", ",")
    return f"{int(round(minutes))} min"


def compute_card_business(db: Session, now: datetime | None = None) -> dict[str, Any]:
    """Numeros de negocio do card semanal: horas economizadas (semana/acumulado), adocao e saude.

    - Horas: arquivos enviados ao workspace x REPORT_MINUTES_PER_FILE (minutos -> horas).
      "Semana" = ultimos 7 dias (janela propria, sempre verdadeira); "Acumulado" = all-time.
    - Engenheiros: solicitantes unicos da lista SharePoint de pedidos de acesso (coluna IDRede,
      trim + case-insensitive), com fallback gracioso para o count local de tarefas
      add_playground_user_to_workspace concluidas se a lista nao puder ser lida.
    - SPECs prontas + saude: reutiliza a classificacao por workspace de block_simplificado (all-time).
    """
    now_utc = now or datetime.utcnow()
    week_start = now_utc - timedelta(days=7)
    minutes_per_file = float(settings.REPORT_MINUTES_PER_FILE or 0)

    sent_filter = (
        WorkspaceFile.is_deleted == False,
        func.lower(WorkspaceFile.status).in_(SENT_FILE_STATUSES),
    )
    files_total = db.query(func.count(WorkspaceFile.id)).filter(*sent_filter).scalar() or 0
    files_week = db.query(func.count(WorkspaceFile.id)).filter(
        *sent_filter,
        func.coalesce(WorkspaceFile.uploaded_at, WorkspaceFile.created_at) >= week_start,
    ).scalar() or 0

    engineers_count = get_engineers_count(db, now=now_utc)

    all_time = {"start": None, "end": None, "automation_id": None, "workspace_id": None, "status": None, "source_task_id": None}
    simpl = block_simplificado(db, all_time, ({}, {}))
    specs_ready = sum(1 for r in simpl.rows if str(r[2]).strip().upper() == "COMPLETO")
    health_items = sum(1 for r in simpl.rows if str(r[2]).strip().upper() == "ERRO")

    return {
        "hours": {
            "week": _format_hours(files_week * minutes_per_file),
            "total": _format_hours(files_total * minutes_per_file),
            "files_week": int(files_week),
            "files_total": int(files_total),
            "minutes_per_file": minutes_per_file,
        },
        "adoption": {"engineers": engineers_count, "specs_ready": specs_ready},
        "health": {"items": health_items},
    }


def compute_card_image_data(db: Session, now: datetime | None = None, language: str = DEFAULT_LANGUAGE) -> dict[str, Any]:
    """Dados dinamicos do poster-convite semanal (a mesma ordem do Adaptive Card de adocao).

    Entrega ao template (report_image.py): 1) convite (manchete/corpo/como pedir acesso);
    2) horas devolvidas ao time (semana + acumulado) + serie cumulativa EM HORAS p/ o grafico;
    3) adocao (engenheiros/SPECs prontas); 4) saude (itens + previsao de correcao).
    FICA DE FORA: contagem de arquivos, tabela SPEC-por-SPEC e status cru de workspace.
    Quando language="en", os textos (manchete/convite/titulo/saude) e os `labels` do poster vao em
    ingles; "pt" (padrao) mantem o comportamento historico intacto.
    """
    language = normalize_language(language)
    cs = card_strings(language)
    ps = poster_strings(language)
    now_utc = now or datetime.utcnow()
    week_start = now_utc - timedelta(days=7)
    business = compute_card_business(db, now_utc)
    hours = business.get("hours", {})
    files_total = int(hours.get("files_total", 0) or 0)
    files_week = int(hours.get("files_week", 0) or 0)
    minutes_per_file = float(hours.get("minutes_per_file", 0) or 0)

    # Serie diaria cumulativa dos ultimos 7 dias, convertida em HORAS devolvidas (prova de valor,
    # nao "arquivos processados"): horas = arquivos_cumulativos x minutos_por_arquivo / 60.
    sent_filter = (
        WorkspaceFile.is_deleted == False,
        func.lower(WorkspaceFile.status).in_(SENT_FILE_STATUSES),
    )
    day_counts: dict[Any, int] = {}
    for (ts,) in db.query(
        func.coalesce(WorkspaceFile.uploaded_at, WorkspaceFile.created_at)
    ).filter(
        *sent_filter,
        func.coalesce(WorkspaceFile.uploaded_at, WorkspaceFile.created_at) >= week_start,
    ).all():
        if ts is None:
            continue
        day = to_sao_paulo_naive(ts, assume_utc=True).date()
        day_counts[day] = day_counts.get(day, 0) + 1
    today_local = to_sao_paulo_naive(now_utc, assume_utc=True).date()
    cumulative = max(0, files_total - files_week)  # base = processados antes da janela de 7 dias
    hours_series: list[dict[str, Any]] = []
    for i in range(6, -1, -1):
        day = today_local - timedelta(days=i)
        cumulative += day_counts.get(day, 0)
        hours_series.append({"label": day.strftime("%d/%m"), "value": round(cumulative * minutes_per_file / 60.0, 1)})

    p_start = to_sao_paulo_naive(week_start, assume_utc=True).strftime("%d/%m/%Y")
    p_end = today_local.strftime("%d/%m/%Y")
    return {
        "brand": ps["brand"],
        "title": ps["title"],
        "period": ps["period_sep"].format(a=p_start, b=p_end),
        "generated_at": to_sao_paulo_naive(now_utc, assume_utc=True).strftime("%d/%m/%Y %H:%M"),
        "headline": cs["headline"],
        "invite_body": cs["invite_body"],
        "access_line": cs["access_line"],
        "language": language,
        "labels": poster_labels(language),
        "playground_url": str(settings.REPORT_CARD_PLAYGROUND_URL or settings.PLAYGROUND_URL or "").strip(),
        "hours": {"week": str(hours.get("week", "0 h")), "total": str(hours.get("total", "0 h"))},
        "hours_series": hours_series,
        "adoption": business.get("adoption", {}),
        "health": {**business.get("health", {}), "eta": cs["health_eta"]},
    }


def build_card_summary(report_type: str, sections: list[ReportSection], rep: ExecutionReport, filters: dict[str, Any], business: dict[str, Any] | None = None, language: str = DEFAULT_LANGUAGE) -> dict[str, Any]:
    """Resumo executivo pronto para o Adaptive Card.

    Para o Relatorio Simplificado com `business`, monta o card de ADOCAO (convite + horas + adocao +
    saude). Demais tipos seguem o resumo classico (metrics + previa). O idioma escolhido fica em
    card["language"] e e lido por build_adaptive_card/build_adoption_card ao renderizar.
    """
    language = normalize_language(language)
    cs = card_strings(language)
    generated_at = fmt_utc(rep.created_at) if rep.created_at else format_dt(now_sao_paulo_naive())
    # Periodo sempre preenchido: se a janela nao veio nos filtros, usa os ultimos 7 dias.
    start, end = filters.get("start"), filters.get("end")
    if not (start and end):
        end = end or rep.created_at or datetime.utcnow()
        start = start or (end - timedelta(days=7))
    card: dict[str, Any] = {
        "title": rep.name or report_type,
        "report_type": report_type,
        "language": language,
        "period": f"{fmt_utc(start)} - {fmt_utc(end)}",
        "generated_at": generated_at,
        "logo_url": str(settings.REPORT_CARD_LOGO_URL or "").strip(),
        "metrics": [],
        "preview": {"headers": [], "rows": []},
    }

    simpl = _section_by_key(sections, "simplificado")
    if simpl is not None and business is not None:
        # Card de adocao: convite como manchete + horas economizadas + adocao + saude em 1 linha.
        card["kind"] = "adoption"
        card["headline"] = cs["headline"]
        card["invite_body"] = cs["invite_body"]
        card["access_line"] = cs["access_line"]
        card["hours"] = business.get("hours", {})
        card["adoption"] = business.get("adoption", {})
        card["health"] = business.get("health", {})
        return card

    if simpl is not None:
        idx = {header: position for position, header in enumerate(simpl.headers)}
        i_spec = idx.get("SPEC", 0)
        i_pct = idx.get("PORCENTAGEM", 1)
        i_status = idx.get("STATUS", 2)
        i_files = idx.get("ARQUIVOS", len(simpl.headers) - 1)

        def cell(row: list[Any], position: int) -> str:
            return normalize_cell(row[position]) if position < len(row) else ""

        def status_of(row: list[Any]) -> str:
            return cell(row, i_status).strip().upper()

        # Comparacoes robustas ao idioma: o valor exibido do STATUS e localizado, entao comparamos
        # com o rotulo localizado do token canonico (em pt == COMPLETO/PROGRESSO/ERRO).
        s_complete = simplificado_status("COMPLETO", language).upper()
        s_progress = simplificado_status("PROGRESSO", language).upper()
        s_error = simplificado_status("ERRO", language).upper()

        total_files = 0
        for row in simpl.rows:
            try:
                total_files += int(str(cell(row, i_files)).strip() or 0)
            except (ValueError, TypeError):
                pass

        card["metrics"] = [
            {"label": cs["m_workspaces"], "value": str(len(simpl.rows))},
            {"label": cs["m_complete"], "value": str(sum(1 for r in simpl.rows if status_of(r) == s_complete))},
            {"label": cs["m_inprogress"], "value": str(sum(1 for r in simpl.rows if status_of(r) == s_progress))},
            {"label": cs["m_error"], "value": str(sum(1 for r in simpl.rows if status_of(r) == s_error))},
            {"label": cs["m_files"], "value": str(total_files)},
        ]
        preview_rows = simpl.rows[:SIMPLIFICADO_PREVIEW_MAX]
        card["preview"] = {
            "headers": ["SPEC", "STATUS", "%"],
            "rows": [[cell(r, i_spec), cell(r, i_status), cell(r, i_pct)] for r in preview_rows],
            "overflow": max(0, len(simpl.rows) - SIMPLIFICADO_PREVIEW_MAX),
        }
    else:
        card["metrics"] = [{"label": section.title, "value": str(len(section.rows))} for section in sections]
        primary = sections[0] if sections else None
        if primary is not None:
            card["preview"] = {
                "headers": list(primary.headers[:3]),
                "rows": [[normalize_cell(value) for value in row[:3]] for row in primary.rows[:CARD_PREVIEW_ROW_LIMIT]],
            }
    return card


def _card_row(cells: list[Any], bold: bool = False) -> dict[str, Any]:
    return {
        "type": "ColumnSet",
        "spacing": "Small",
        "columns": [
            {
                "type": "Column",
                "width": "stretch",
                "items": [{
                    "type": "TextBlock",
                    "text": str(cell),
                    "wrap": True,
                    "size": "Small",
                    "weight": "Bolder" if bold else "Default",
                }],
            }
            for cell in cells
        ],
    }


def _adoption_header(card: dict[str, Any]) -> list[dict[str, Any]]:
    brand = {"type": "TextBlock", "text": "Stellantis Automation HUB", "weight": "Bolder", "size": "Small", "color": "Accent", "spacing": "None"}
    logo_url = str(card.get("logo_url") or "").strip()
    if logo_url:
        return [{
            "type": "ColumnSet",
            "columns": [
                {"type": "Column", "width": "auto", "verticalContentAlignment": "Center",
                 "items": [{"type": "Image", "url": logo_url, "height": "40px", "altText": "Stellantis"}]},
                {"type": "Column", "width": "stretch", "items": [brand]},
            ],
        }]
    return [brand]


def build_adoption_card(card: dict[str, Any]) -> dict[str, Any]:
    """Card semanal de ADOCAO, na ordem do convite: manchete -> horas -> adocao -> saude.

    1) convite (manchete + corpo + como pedir acesso); 2) horas devolvidas; 3) adocao; 4) saude em
    1 linha; rodape com periodo/gerado e os botoes.
    """
    language = normalize_language(card.get("language"))
    cs = card_strings(language)
    body: list[dict[str, Any]] = _adoption_header(card)

    # 1) Convite: manchete + corpo + "como pedir acesso".
    body.append({"type": "TextBlock", "text": str(card.get("headline", "")), "weight": "Bolder", "size": "Large", "wrap": True, "spacing": "Small"})
    if card.get("invite_body"):
        body.append({"type": "TextBlock", "text": str(card["invite_body"]), "wrap": True, "spacing": "Small"})
    if card.get("access_line"):
        body.append({"type": "TextBlock", "text": str(card["access_line"]), "wrap": True, "isSubtle": True, "spacing": "Small"})

    # 2) Horas devolvidas ao time (prova de valor).
    hours = card.get("hours") or {}
    minutes_per_file = float(hours.get("minutes_per_file", 4) or 0)
    body.append({"type": "TextBlock", "text": cs["hours_title"], "weight": "Bolder", "separator": True, "spacing": "Medium", "wrap": True})
    body.append({"type": "FactSet", "facts": [
        {"title": cs["this_week"], "value": str(hours.get("week", "0 min"))},
        {"title": cs["total"], "value": str(hours.get("total", "0 min"))},
    ]})
    body.append({"type": "TextBlock", "text": cs["hours_note"].format(mpf=minutes_per_file), "isSubtle": True, "wrap": True, "spacing": "None", "size": "Small"})

    # 3) Adocao.
    adoption = card.get("adoption") or {}
    body.append({"type": "TextBlock", "text": cs["adoption_title"], "weight": "Bolder", "separator": True, "spacing": "Medium", "wrap": True})
    body.append({"type": "FactSet", "facts": [
        {"title": cs["engineers"], "value": str(adoption.get("engineers", 0))},
        {"title": cs["specs_ready"], "value": str(adoption.get("specs_ready", 0))},
    ]})

    # 4) Saude em 1 linha: itens em tratamento + previsao de correcao.
    items = int((card.get("health") or {}).get("items") or 0)
    if items > 0:
        eta = str((card.get("health") or {}).get("eta") or cs["health_eta"]).strip()
        eta_txt = cs["health_eta_prefix"].format(eta=eta) if eta else ""
        body.append({"type": "TextBlock", "text": cs["health_warn"].format(items=items, eta=eta_txt), "color": "Warning", "wrap": True, "separator": True, "spacing": "Medium"})
    else:
        body.append({"type": "TextBlock", "text": cs["health_ok"], "color": "Good", "wrap": True, "separator": True, "spacing": "Medium"})

    # Rodape: periodo + gerado em (discretos).
    body.append({"type": "TextBlock", "text": cs["period"].format(v=card.get('period', '')), "isSubtle": True, "wrap": True, "spacing": "Small", "size": "Small"})
    body.append({"type": "TextBlock", "text": cs["generated"].format(v=card.get('generated_at', '')), "isSubtle": True, "wrap": True, "spacing": "None", "size": "Small"})

    actions: list[dict[str, Any]] = []
    if str(card.get("access_url") or "").strip():
        actions.append(
            build_access_request_showcard(
                language,
                title_override="Solicitar acesso" if language == "pt" else None,
            )
        )
    actions.append({"type": "Action.OpenUrl", "title": cs["view_details_pdf"], "url": DOWNLOAD_URL_PLACEHOLDER})

    return {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "msteams": {"width": "Full"},
        "body": body,
        "actions": actions,
    }


def build_access_request_showcard(
    language: str = DEFAULT_LANGUAGE,
    *,
    title_override: str | None = None,
) -> dict[str, Any]:
    """Sub-card (Action.ShowCard) do formulario de solicitacao de acesso, revelado inline no
    proprio card semanal — sem navegar para fora do Teams e sem precisar de um 2o fluxo/Run.

    O Action.Submit devolve `data.acao = "solicitar_acesso"` + idrede/spec/justificativa; o fluxo
    do Power Automate que postou o card (via "Post adaptive card and wait for a response") recebe
    a resposta e quem enviou (`responder`), grava na Lista do SharePoint e avisa o aprovador.
    """
    cs = card_strings(language)
    return {
        "type": "Action.ShowCard",
        "title": title_override or cs["request_access"],
        "card": {
            "type": "AdaptiveCard",
            "body": [
                {"type": "TextBlock", "text": cs["sc_title"], "weight": "Bolder", "wrap": True},
                {"type": "TextBlock", "text": cs["sc_subtitle"], "isSubtle": True, "wrap": True, "spacing": "None"},
                {"type": "Input.Text", "id": "idrede", "label": cs["sc_idrede_label"], "placeholder": cs["sc_idrede_ph"], "isRequired": True, "errorMessage": cs["sc_idrede_err"]},
                {"type": "Input.Text", "id": "spec", "label": cs["sc_spec_label"], "placeholder": cs["sc_spec_ph"], "isRequired": True, "errorMessage": cs["sc_spec_err"]},
                {"type": "Input.Text", "id": "justificativa", "label": cs["sc_just_label"], "placeholder": cs["sc_just_ph"], "isMultiline": True},
            ],
            "actions": [
                {"type": "Action.Submit", "title": cs["sc_submit"], "data": {"acao": "solicitar_acesso"}},
            ],
        },
    }


# NOTA (image_placeholder/download_placeholder abaixo): quando REPORT_BACKEND_BASE_URL esta
# configurada, write_report_to_delivery_folder substitui os dois placeholders pelos links DIRETOS
# do backend antes de gravar o sidecar -- o Power Automate deixa de precisar trocar esses
# placeholders pelos links do OneDrive (menos passos, menos pontos de falha).
def build_report_image_card(
    image_placeholder: str = IMAGE_URL_PLACEHOLDER,
    download_placeholder: str = DOWNLOAD_URL_PLACEHOLDER,
    language: str = DEFAULT_LANGUAGE,
) -> dict[str, Any]:
    """Card-imagem do Teams: o PNG fiel do relatorio (via Image) + botoes de acao.

    Ordem dos botoes: Abrir Playground (CTA principal) -> Solicitar Acesso (ShowCard inline,
    formulario revelado dentro do proprio card) -> Baixar Relatorio (PDF). O placeholder de
    imagem/PDF e trocado pelo Power Automate (ou, se REPORT_BACKEND_BASE_URL estiver configurada,
    ja chega substituido pelo link direto do backend -- ver write_report_to_delivery_folder).

    NOTA (10/07/2026): cheguei a trocar "Solicitar Acesso" para Action.OpenUrl aqui, achando que o
    ShowCard nao funcionava (a documentacao generica do GUIA_POWER_AUTOMATE.md posta o card com
    "Post card in a chat or channel", que de fato nao processa Action.Submit). Revertido: o fluxo
    REAL no Power Automate ("HUB - Relatorios Teams") posta este card com "Post adaptive card and
    wait for a response" seguido de "Create item" (SharePoint) + "Post message in a chat or
    channel" -- ou seja, o ShowCard FOI PROJETADO para funcionar nesse fluxo especifico (o
    Action.Submit do formulario embutido volta como resposta para essa mesma acao). Mantido como
    estava.
    """
    cs = card_strings(language)
    playground_url = str(settings.REPORT_CARD_PLAYGROUND_URL or settings.PLAYGROUND_URL or "").strip()
    access_url = str(settings.REPORT_CARD_ACCESS_URL or "").strip()
    actions: list[dict[str, Any]] = []
    if playground_url:
        actions.append({"type": "Action.OpenUrl", "title": cs["open_playground"], "url": playground_url})
    if access_url:
        actions.append(build_access_request_showcard(language))
    actions.append({"type": "Action.OpenUrl", "title": cs["download_pdf"], "url": download_placeholder})
    return {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "msteams": {"width": "Full"},
        "body": [{
            "type": "Image",
            "url": image_placeholder,
            "size": "Stretch",
            "altText": "Relatório Semanal — Stellantis Automation HUB",
        }],
        "actions": actions,
    }


def build_adaptive_card(card: dict[str, Any]) -> dict[str, Any]:
    """Adaptive Card 1.4 (compativel com Teams) pronto para postar verbatim no Power Automate."""
    if card.get("kind") == "adoption":
        return build_adoption_card(card)
    language = normalize_language(card.get("language"))
    cs = card_strings(language)
    brand = {"type": "TextBlock", "text": cs["brand"], "weight": "Bolder", "size": "Small", "color": "Accent", "spacing": "None"}
    title = {"type": "TextBlock", "text": card.get("title", ""), "weight": "Bolder", "size": "Large", "wrap": True}
    logo_url = str(card.get("logo_url") or "").strip()
    if logo_url:
        # Logo (URL publica) a esquerda + marca/titulo a direita.
        header: dict[str, Any] = {
            "type": "ColumnSet",
            "columns": [
                {
                    "type": "Column",
                    "width": "auto",
                    "verticalContentAlignment": "Center",
                    "items": [{"type": "Image", "url": logo_url, "height": "40px", "altText": "Stellantis"}],
                },
                {"type": "Column", "width": "stretch", "items": [brand, title]},
            ],
        }
        body: list[dict[str, Any]] = [header]
    else:
        body = [brand, title]
    body.extend([
        {"type": "TextBlock", "text": cs["period_classic"].format(v=card.get('period', '')), "isSubtle": True, "wrap": True, "spacing": "None"},
        {"type": "TextBlock", "text": cs["generated_classic"].format(v=card.get('generated_at', '')), "isSubtle": True, "wrap": True, "spacing": "None"},
    ])
    metrics = card.get("metrics") or []
    if metrics:
        body.append({
            "type": "FactSet",
            "separator": True,
            "spacing": "Medium",
            "facts": [{"title": str(m.get("label", "")), "value": str(m.get("value", ""))} for m in metrics],
        })
    preview = card.get("preview") or {}
    preview_rows = preview.get("rows") or []
    if preview_rows:
        body.append({"type": "TextBlock", "text": cs["preview"].format(type=card.get('report_type', '')), "weight": "Bolder", "separator": True, "spacing": "Medium", "wrap": True})
        headers = preview.get("headers") or []
        if headers:
            body.append(_card_row(headers, bold=True))
        for row in preview_rows:
            body.append(_card_row(row))
        overflow = int(preview.get("overflow") or 0)
        if overflow > 0:
            body.append({"type": "TextBlock", "text": cs["overflow"].format(n=overflow), "isSubtle": True, "wrap": True, "spacing": "Small"})
    return {
        "type": "AdaptiveCard",
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.4",
        "msteams": {"width": "Full"},
        "body": body,
        "actions": [
            {"type": "Action.OpenUrl", "title": cs["download_pdf_short"], "url": DOWNLOAD_URL_PLACEHOLDER},
        ],
    }


def report_delivery_bundle(db: Session, report_id: int) -> dict[str, Any]:
    rep = db.query(ExecutionReport).filter(
        ExecutionReport.id == report_id,
        ExecutionReport.is_deleted == False,
        ExecutionReport.source_scope == REPORT_SOURCE_SCOPE,
    ).first()
    if not rep:
        raise HTTPException(404, detail="Relatorio nao encontrado.")
    language = normalize_language(getattr(rep, "language", None))
    report_type, file_format = report_type_and_format(rep)
    if rep.file_path and Path(rep.file_path).exists():
        content = Path(rep.file_path).read_bytes()
    else:
        content = fallback_content(rep, file_format, db)
    filename = Path(rep.file_path).name if rep.file_path else f"{clean_filename(rep.name or f'report_{rep.id}')}.{file_format}"
    media_type = MEDIA_TYPES.get(file_format, MEDIA_TYPES["csv"])
    summary = (
        f"Relatorio: {rep.name or report_type}\n"
        f"Tipo: {report_type}\n"
        f"Periodo: {fmt_utc(rep.period_start)} - {fmt_utc(rep.period_end)}\n"
        f"Gerado em: {fmt_utc(rep.created_at)}"
    )
    # Conteudo estruturado para o card (resumo + previa) e PDF companheiro para anexo no Teams.
    filters = filters_for_report(rep)
    sections = build_sections(db, report_type, filters, language)
    business = compute_card_business(db)
    card = build_card_summary(report_type, sections, rep, filters, business=business, language=language)
    if card.get("kind") == "adoption":
        card["access_url"] = str(settings.REPORT_CARD_ACCESS_URL or "").strip()
    adaptive_card = build_adaptive_card(card)
    # Card-imagem (PNG fiel ao mockup) so no relatorio de adocao/simplificado (o do card semanal).
    image_data = compute_card_image_data(db, language=language) if card.get("kind") == "adoption" else None
    if file_format == "pdf":
        pdf_content = content
        pdf_filename = filename
    else:
        pdf_content = build_report_content(report_type, "pdf", filters, db, language)
        pdf_filename = f"{Path(filename).stem}.pdf"
    return {
        "report": rep,
        "report_type": report_type,
        "file_format": file_format,
        "content": content,
        "filename": filename,
        "media_type": media_type,
        "summary": summary,
        "sections": sections,
        "card": card,
        "adaptive_card": adaptive_card,
        "image_data": image_data,
        "pdf_content": pdf_content,
        "pdf_filename": pdf_filename,
    }


def _render_report_image_threaded(image_data: dict[str, Any], out_path: Path, timeout_s: float = 45.0) -> Path | None:
    """Renderiza o PNG numa thread propria e retorna o Path (ou None).

    A Sync API do Playwright nao roda dentro de um event loop asyncio; uma thread nova nao tem
    loop em execucao, entao isto funciona igual em endpoint async, no scheduler ou em scripts sync.
    """
    import threading

    from app.services.report_image import generate_report_image

    result: dict[str, Path | None] = {"path": None}

    def _job() -> None:
        try:
            result["path"] = generate_report_image(image_data, out_path)
        except Exception:
            result["path"] = None

    thread = threading.Thread(target=_job, name="report-image-render", daemon=True)
    thread.start()
    thread.join(timeout=timeout_s)
    return result["path"]


def write_report_to_delivery_folder(bundle: dict, routing: dict | None = None, environment: str | None = None) -> Path | None:
    target = report_delivery_dir(environment)
    if target is None:
        return None
    target.mkdir(parents=True, exist_ok=True)
    report_path = target / bundle["filename"]
    if report_path.exists():
        stem, suffix = report_path.stem, report_path.suffix
        counter = 1
        while report_path.exists():
            report_path = target / f"{stem}_{counter}{suffix}"
            counter += 1
    report_path.write_bytes(bundle["content"])
    rep = bundle["report"]

    # PDF companheiro (anexo do Teams). Se o relatorio ja for PDF, ele proprio e o anexo;
    # caso contrario gravamos "{stem}.pdf" ao lado, junto com os lotes/arquivos.
    pdf_content = bundle.get("pdf_content")
    if report_path.suffix.lower() == ".pdf":
        attachment_path = report_path
    elif pdf_content:
        attachment_path = target / f"{report_path.stem}.pdf"
        attachment_path.write_bytes(pdf_content)
    else:
        attachment_path = report_path

    # Card-imagem (PNG fiel ao mockup): renderiza so quando ha image_data (relatorio de adocao) e
    # o Chromium offline consegue gerar. Falhou (sem browser/timeout)? cai no card-texto (fallback),
    # sem quebrar a entrega.
    language = normalize_language(getattr(rep, "language", None))
    image_data = bundle.get("image_data")
    image_path = None
    if image_data:
        image_path = _render_report_image_threaded(image_data, target / f"{report_path.stem}.png")
    if image_path is not None:
        adaptive_card = build_report_image_card(language=language)
    else:
        adaptive_card = bundle.get("adaptive_card")

    sidecar = {
        "report_id": rep.id,
        "name": rep.name,
        "report_type": bundle["report_type"],
        "file_format": bundle["file_format"],
        "period_start": sao_paulo_utc_iso(rep.period_start),
        "period_end": sao_paulo_utc_iso(rep.period_end),
        "generated_at": sao_paulo_utc_iso(rep.created_at),
        "report_file": report_path.name,
        "attachment_file": attachment_path.name,
        "download_url_placeholder": DOWNLOAD_URL_PLACEHOLDER,
        "card": bundle.get("card"),
        "adaptive_card": adaptive_card,
    }
    if image_path is not None:
        sidecar["image_file"] = image_path.name
        sidecar["image_url_placeholder"] = IMAGE_URL_PLACEHOLDER

    # Links DIRETOS do proprio backend (opcional, ver REPORT_BACKEND_BASE_URL em config.py).
    # Quando configurada, ja substituimos os placeholders AQUI (antes de gravar o sidecar) pelos
    # links estaveis do backend -- o Power Automate posta 'adaptive_card' verbatim, sem precisar
    # dos passos "Create share link" do OneDrive (Parte I.4/I.5 do GUIA_POWER_AUTOMATE.md).
    base_url = str(settings.REPORT_BACKEND_BASE_URL or "").strip()
    if base_url:
        base_url = base_url.rstrip("/")
        download_direct_url = f"{base_url}/api/reports/{rep.id}/download"
        sidecar["download_direct_url"] = download_direct_url
        card_text = json.dumps(sidecar["adaptive_card"], ensure_ascii=False)
        card_text = card_text.replace(DOWNLOAD_URL_PLACEHOLDER, download_direct_url)
        if image_path is not None:
            image_direct_url = f"{base_url}/api/reports/{rep.id}/image"
            sidecar["image_direct_url"] = image_direct_url
            card_text = card_text.replace(IMAGE_URL_PLACEHOLDER, image_direct_url)
        sidecar["adaptive_card"] = json.loads(card_text)

    for key in ("teams_channel", "email_to", "subject"):
        if routing and routing.get(key):
            sidecar[key] = routing[key]
    # Sidecar gravado POR ULTIMO (gatilho seguro do flow: relatorio e PDF ja existem) e SEMPRE
    # com sufixo ".meta.json". Assim "*.meta.json" e inequivocamente o sidecar para qualquer
    # formato -- inclusive .json, cujo proprio arquivo do relatorio nao pode ser sobrescrito.
    sidecar_path = target / f"{report_path.stem}.meta.json"
    sidecar_path.write_text(json.dumps(sidecar, ensure_ascii=False, default=str), encoding="utf-8")
    return report_path


@router.get("")
def list_reports(limit: int = Query(100, ge=1, le=500), db: Session = Depends(get_db)):
    reports = (
        db.query(ExecutionReport)
        .filter(ExecutionReport.is_deleted == False, ExecutionReport.source_scope == REPORT_SOURCE_SCOPE)
        .order_by(ExecutionReport.created_at.desc(), ExecutionReport.id.desc())
        .limit(limit)
        .all()
    )
    return [report_out(report) for report in reports]


@router.post("")
def create_report(data: dict, db: Session = Depends(get_db)):
    report_type = parse_report_type(data.get("report_type"))
    file_format = parse_file_format(data.get("file_format"))
    if file_format not in MEDIA_TYPES:
        raise HTTPException(422, detail="Formato de relatorio invalido.")
    filters = filters_from_payload(data)
    # Idioma opcional do relatorio ("pt" padrao ou "en"); omitido -> pt (comportamento historico).
    language = normalize_language(data.get("language"))
    environment_mode = parse_environment_mode(data.get("environment_mode") or data.get("app_mode"))
    if environment_mode == "developer":
        build_report_content(report_type, file_format, filters, db, language)
        return {
            "report": None,
            "saved": False,
            "environment_mode": "developer",
            "language": language,
            "message": "Modo Desenvolvedor: relatorio processado para teste e nada foi salvo.",
        }
    report = persist_report(
        db,
        report_type,
        file_format,
        filters,
        safe_int(data.get("generated_by_id") or data.get("generated_by_user_id")),
        "manual",
        None,
        language=language,
    )
    return {"report": report_out(report), "saved": True, "environment_mode": "operational"}


@router.get("/{id}/download")
def download_report(id: int, language: str | None = Query(None), db: Session = Depends(get_db)):
    report = db.query(ExecutionReport).filter(
        ExecutionReport.id == id,
        ExecutionReport.is_deleted == False,
        ExecutionReport.source_scope == REPORT_SOURCE_SCOPE,
    ).first()
    if not report:
        raise HTTPException(404, detail="Relatorio nao encontrado.")
    _, file_format = report_type_and_format(report)
    safe_name = Path(report.file_path).name if report.file_path else f"{clean_filename(report.name or f'report_{report.id}')}.{file_format}"
    media_type = MEDIA_TYPES.get(file_format, MEDIA_TYPES["csv"])
    stored_language = normalize_language(getattr(report, "language", None))
    # `language` na query permite baixar em outro idioma (regenera o conteudo); sem ela, serve o
    # arquivo salvo (ou regenera no idioma persistido, se o arquivo sumiu).
    override = language is not None and normalize_language(language) != stored_language
    if report.file_path and Path(report.file_path).exists() and not override:
        create_log(db, "info", "Downloaded report", "report", report.id, metadata={"source_scope": REPORT_SOURCE_SCOPE})
        return FileResponse(report.file_path, media_type=media_type, filename=safe_name)
    effective_language = normalize_language(language) if language is not None else stored_language
    content = build_report_content(*report_type_and_format(report), filters_for_report(report), db, effective_language) if override else fallback_content(report, file_format, db)
    create_log(db, "info", "Downloaded report", "report", report.id, metadata={"source_scope": REPORT_SOURCE_SCOPE, "language": effective_language})
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@router.get("/{id}/image")
def download_report_image(id: int, language: str | None = Query(None), db: Session = Depends(get_db)):
    """PNG do card semanal (poster-convite), gerado NA HORA e servido direto pelo backend.

    Pensado para o `Image.url` do Adaptive Card do Teams: um link ESTAVEL (nao expira, nao depende
    de politica de compartilhamento) -- alternativa ao link do OneDrive (GUIA_POWER_AUTOMATE.md,
    Parte I), que e o ponto mais fragil da entrega hoje (comportamento nao documentado do
    `&download=1`). So funciona para relatorios com card de adocao (Relatorio Simplificado); demais
    tipos nao tem imagem (404).
    """
    report = db.query(ExecutionReport).filter(
        ExecutionReport.id == id,
        ExecutionReport.is_deleted == False,
        ExecutionReport.source_scope == REPORT_SOURCE_SCOPE,
    ).first()
    if not report:
        raise HTTPException(404, detail="Relatorio nao encontrado.")

    effective_language = normalize_language(language) if language is not None else normalize_language(getattr(report, "language", None))
    image_data = compute_card_image_data(db, language=effective_language)
    import tempfile

    tmp_dir = Path(tempfile.mkdtemp(prefix="report_image_"))
    image_path = _render_report_image_threaded(image_data, tmp_dir / f"report_{id}.png")
    if image_path is None:
        raise HTTPException(502, detail="Nao foi possivel gerar a imagem do relatorio (Chromium/Playwright offline indisponivel).")

    create_log(db, "info", "Rendered report image", "report", report.id, metadata={"source_scope": REPORT_SOURCE_SCOPE})
    return StreamingResponse(io.BytesIO(image_path.read_bytes()), media_type="image/png")


@router.delete("/{id}")
def delete_report(id: int, db: Session = Depends(get_db)):
    report = db.query(ExecutionReport).filter(
        ExecutionReport.id == id,
        ExecutionReport.is_deleted == False,
        ExecutionReport.source_scope == REPORT_SOURCE_SCOPE,
    ).first()
    if not report:
        raise HTTPException(404, detail="Report not found")
    report.is_deleted = True
    report.deleted_at = datetime.utcnow()
    db.commit()
    create_log(db, "warning", f"Report marked as deleted: {report.name}", "report", report.id)
    return {"status": "deleted"}
